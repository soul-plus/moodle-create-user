from openpyxl import load_workbook, Workbook
from cyrtranslit import to_latin
from tkinter import filedialog
import tkinter
import secrets
import string
import os


def transliterate(text):
    return to_latin(text, 'ru')


def get_username(text, current_year):
    name = transliterate(text)
    name_to_list = name.split()
    if len(name_to_list) >= 3:
        username = name_to_list[0].lower() + name_to_list[1][0].lower() + name_to_list[2][0].lower() + str(current_year)
    else:
        username = name_to_list[0].lower() + str(current_year)
    username = str.replace(username, '\'', '')
    return username


def create_password():
    alphabet = string.ascii_letters + string.digits + '#@!'
    alphabet = str.replace(alphabet, 'I', '')
    alphabet = str.replace(alphabet, 'l', '')
    while True:
        password = ''.join(secrets.choice(alphabet) for i in range(10))
        if (any(c.islower() for c in password)
                and any(c.isupper() for c in password)
                and sum(c.isdigit() for c in password) >= 1
                and any(c in '#!' for c in password)):
            break
    return password


def validate(work_sheet):
    if work_sheet[1][0].value == 'fio' and work_sheet[1][1].value == 'email' and work_sheet[1][2].value == 'group':
        return 'npk'
    else:
        if work_sheet[1][0].value == 'fio' and work_sheet[1][1].value == 'group':
            return 'stud'
        else:
            return False


def get_file_name(file_path, file):
    if os.path.isfile(os.path.join(file_path, file)):
        i = 1
        temp_name, extension = str.rsplit(file, '.', 1)
        while True:
            file = temp_name + '(' + str(i) + ').' + extension
            i += 1
            if not os.path.isfile(os.path.join(file_path, file)):
                break
        return file
    else:
        return file


def create_new_workbook(work_sheet, result_sheet, group_prefix, current_year):
    sheet_type = validate(work_sheet)
    group_is_set = False
    email_is_set = False
    group_input = ''

    if sheet_type == 'npk':
        email_is_set = work_sheet[2][1].value is not None
        group_is_set = work_sheet[2][2].value is not None
    elif sheet_type == 'stud':
        group_is_set = work_sheet[2][1].value is not None

    name_list = ('username', 'password', 'firstname', 'lastname', 'cohort1', 'email')

    if not group_is_set:
        group_input = input('Введите номер группы в moodle: \n')

    for i in range(0, 6):
        final_sheet.cell(1, i + 1).value = name_list[i]
    for i in range(2, work_sheet.max_row + 1):
        fio = work_sheet[i][0].value
        username = get_username(fio, current_year)
        result_sheet.cell(i, 1).value = username
        result_sheet.cell(i, 2).value = create_password()
        result_sheet.cell(i, 3).value = fio
        if group_is_set and sheet_type == 'npk':
            group = group_prefix + str(work_sheet[i][2].value)
        elif group_is_set and sheet_type == 'stud':
            group = group_prefix + str(work_sheet[i][1].value)
        else:
            group = group_input
        result_sheet.cell(i, 4).value = group
        result_sheet.cell(i, 5).value = group
        if email_is_set:
            result_sheet.cell(i, 6).value = work_sheet[i][1].value
        else:
            result_sheet.cell(i, 6).value = username + '@kazgik.ru'
    return final_workbook


if __name__ == '__main__':
    year = 23
    prefix = ''

    try:
        tkinter.Tk().withdraw()
        path_to_file = tkinter.filedialog.askopenfilename()
        book = load_workbook(filename=path_to_file)
        worksheets = book.worksheets
        sheet = worksheets[0]

        temp_path = path_to_file.rsplit('/', 1)
        dir_path = temp_path[0] + '/done/'
        file_name = temp_path[1]

        if not os.path.exists(dir_path):
            os.mkdir(dir_path)

        file_name = get_file_name(dir_path, file_name)
        final_path = os.path.join(dir_path, file_name)
        final_workbook = Workbook()
        final_sheet = final_workbook.active

        if validate(sheet) is False:
            print('\033[91m' + 'Validation failed. Template is not correct')
        else:
            create_new_workbook(sheet, final_sheet, prefix, year)
            final_workbook.save(str(final_path))
            print('\033[32m' + 'OK. Saved to: ', final_path)
    except Exception as e:
        print(e)
